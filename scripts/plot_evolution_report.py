"""
Gera PNG e Excel com gráficos a partir de evolution_report.csv.

Uso
---
  python scripts/plot_evolution_report.py
  python scripts/plot_evolution_report.py --input evolution_report.csv --png out.png --xlsx out.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


def load_report(input_path: Path) -> pd.DataFrame:
    df = pd.read_csv(input_path, parse_dates=["ref_date"])
    required = {"ref_date", "mode", "n_train", "n_eval", "n_days_in_month", "rmse", "mae"}
    missing = required - set(df.columns)
    if missing:
        names = ", ".join(sorted(missing))
        raise ValueError(f"Colunas ausentes em {input_path}: {names}")
    return df.sort_values(["mode", "ref_date"]).reset_index(drop=True)


def _mode_subset(df: pd.DataFrame, mode: str) -> pd.DataFrame:
    subset = df.loc[df["mode"] == mode].copy()
    if subset.empty:
        raise ValueError(f"Nenhuma linha com mode={mode!r} no CSV.")
    return subset


def plot_png(df: pd.DataFrame, output_path: Path) -> None:
    pipeline = _mode_subset(df, "pipeline")
    walk_forward = _mode_subset(df, "walk_forward")

    fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=False)
    fig.suptitle("Evolucao mensal do modelo — Bike Sharing (2011–2012)", fontsize=14, y=0.98)

    year_change = pd.Timestamp("2012-01-01")

    for ax, subset, title in (
        (axes[0], pipeline, "Modo pipeline (esteira AWS — treino so no mes)"),
        (axes[1], walk_forward, "Modo walk_forward (historico acumulado → mes inteiro)"),
    ):
        ax.plot(subset["ref_date"], subset["mae"], marker="o", label="MAE", color="#2563eb", linewidth=2)
        ax.plot(subset["ref_date"], subset["rmse"], marker="s", label="RMSE", color="#dc2626", linewidth=2)
        ax.axvline(year_change, color="#6b7280", linestyle="--", linewidth=1, alpha=0.8, label="Inicio 2012")
        ax.set_title(title, fontsize=11, loc="left")
        ax.set_ylabel("Erro (bicicletas/dia)")
        ax.grid(True, alpha=0.3)
        ax.legend(loc="upper left", fontsize=9)
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
        plt.setp(ax.get_xticklabels(), rotation=45, ha="right")

    axes[1].set_xlabel("Mes (ref_date)")
    fig.tight_layout(rect=[0, 0, 1, 0.96])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def _add_line_chart(
    ws,
    *,
    title: str,
    categories_col: int,
    value_cols: list[tuple[int, str]],
    anchor: str,
    width: float = 18,
    height: float = 10,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.style = 10
    chart.y_axis.title = "Erro (bicicletas/dia)"
    chart.x_axis.title = "Mes"
    chart.width = width
    chart.height = height

    max_row = ws.max_row
    cats = Reference(ws, min_col=categories_col, min_row=2, max_row=max_row)
    chart.set_categories(cats)

    for col_idx, _ in value_cols:
        values = Reference(ws, min_col=col_idx, min_row=1, max_row=max_row)
        chart.add_data(values, titles_from_data=True)

    ws.add_chart(chart, anchor)


def _write_mode_sheet(wb: Workbook, sheet_name: str, subset: pd.DataFrame) -> None:
    export = subset.copy()
    export["ref_date"] = export["ref_date"].dt.strftime("%Y-%m-%d")
    export["mae"] = export["mae"].round(2)
    export["rmse"] = export["rmse"].round(2)

    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(export, index=False, header=True):
        ws.append(row)

    _add_line_chart(
        ws,
        title=f"MAE e RMSE — {sheet_name}",
        categories_col=1,
        value_cols=[(6, "rmse"), (7, "mae")],
        anchor="J2",
    )

    if "n_train" in export.columns:
        train_chart = LineChart()
        train_chart.title = "Dias de treino (n_train)"
        train_chart.style = 10
        train_chart.y_axis.title = "Dias"
        train_chart.width = 18
        train_chart.height = 8
        max_row = ws.max_row
        train_chart.add_data(
            Reference(ws, min_col=3, min_row=1, max_row=max_row),
            titles_from_data=True,
        )
        train_chart.set_categories(
            Reference(ws, min_col=1, min_row=2, max_row=max_row),
        )
        ws.add_chart(train_chart, "J20")


def write_xlsx(df: pd.DataFrame, output_path: Path) -> None:
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Dados"

    export_all = df.copy()
    export_all["ref_date"] = export_all["ref_date"].dt.strftime("%Y-%m-%d")
    export_all["mae"] = export_all["mae"].round(2)
    export_all["rmse"] = export_all["rmse"].round(2)
    for row in dataframe_to_rows(export_all, index=False, header=True):
        ws_data.append(row)

    _write_mode_sheet(wb, "Pipeline", _mode_subset(df, "pipeline"))
    _write_mode_sheet(wb, "Walk_forward", _mode_subset(df, "walk_forward"))

    resumo = df.groupby("mode")[["rmse", "mae"]].agg(["mean", "median"]).round(2)
    resumo.columns = [f"{metric}_{stat}" for metric, stat in resumo.columns]
    resumo = resumo.reset_index()

    ws_resumo = wb.create_sheet("Resumo")
    for row in dataframe_to_rows(resumo, index=False, header=True):
        ws_resumo.append(row)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Gera PNG e Excel a partir de evolution_report.csv.")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("evolution_report.csv"),
        help="CSV gerado por simulate_monthly_evolution.py",
    )
    parser.add_argument(
        "--png",
        type=Path,
        default=Path("evolution_report.png"),
        help="Saida PNG",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=Path("evolution_report.xlsx"),
        help="Saida Excel com graficos",
    )
    parser.add_argument("--no-png", action="store_true", help="Nao gerar PNG")
    parser.add_argument("--no-xlsx", action="store_true", help="Nao gerar Excel")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)

    if not args.input.is_file():
        print(f"Arquivo nao encontrado: {args.input}", file=sys.stderr)
        print("Execute antes: python scripts/simulate_monthly_evolution.py ...", file=sys.stderr)
        return 1

    df = load_report(args.input)

    if not args.no_png:
        plot_png(df, args.png)
        print(f"PNG: {args.png.resolve()}")

    if not args.no_xlsx:
        write_xlsx(df, args.xlsx)
        print(f"Excel: {args.xlsx.resolve()}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
